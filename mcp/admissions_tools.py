"""
招生 MCP 工具：录取风险评估。

这个模块承接结构化、可计算的数据能力：
  - 历年各省各专业最低分、最低位次、计划数
  - 根据考生分数/位次输出“冲/稳/保/信息不足”的倾向判断

RAG 知识库适合回答招生章程、专业介绍、校园生活等文本资料；
本工具适合处理“我河南理科 580 分、位次 32000，报计算机稳吗？”这类需要
结构化比较和确定性计算的问题。
"""
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class AdmissionRecord:
    year: int
    province: str
    subject_type: str
    major: str
    min_score: int
    min_rank: int
    plan: Optional[int] = None


# # 演示用 mock 数据。生产环境建议替换为数据库、Excel、官网接口或数据中台查询。
# _ADMISSION_RECORDS: List[AdmissionRecord] = [
#     AdmissionRecord(2025, "河南", "理科", "计算机科学与技术", 575, 35000, 20),
#     AdmissionRecord(2024, "河南", "理科", "计算机科学与技术", 578, 33000, 18),
#     AdmissionRecord(2023, "河南", "理科", "计算机科学与技术", 570, 36000, 22),
#     AdmissionRecord(2025, "河南", "理科", "软件工程", 572, 36500, 24),
#     AdmissionRecord(2024, "河南", "理科", "软件工程", 574, 35000, 20),
#     AdmissionRecord(2023, "河南", "理科", "软件工程", 568, 38000, 26),
#     AdmissionRecord(2025, "河北", "物理类", "计算机科学与技术", 590, 21000, 35),
#     AdmissionRecord(2024, "河北", "物理类", "计算机科学与技术", 588, 22000, 32),
#     AdmissionRecord(2023, "河北", "物理类", "计算机科学与技术", 585, 23500, 34),
#     AdmissionRecord(2025, "天津", "综合改革", "计算机科学与技术", 625, 7800, 30),
#     AdmissionRecord(2024, "天津", "综合改革", "计算机科学与技术", 622, 8100, 28),
#     AdmissionRecord(2023, "天津", "综合改革", "计算机科学与技术", 620, 8400, 30),
# ]
_ADMISSION_SOURCES = [
    {
        "province": "河北",
        "default_subject_type": None,
        "path_or_url": "data/hebut_hebei.xlsx",
    },
    {
        "province": "天津",
        "default_subject_type": "综合改革",
        "path_or_url": "data/hebut_tianjin.xlsx",
    },
]

_SUPPORTED_PROVINCES = {str(source["province"]) for source in _ADMISSION_SOURCES}


def load_admission_records(sources: List[Dict[str, Any]]) -> List[AdmissionRecord]:
    """从 data 下的 Excel 文件加载真实历年录取数据。"""
    records: List[AdmissionRecord] = []
    for source in sources:
        province = str(source["province"])
        default_subject_type = source.get("default_subject_type")
        path = _resolve_data_path(str(source["path_or_url"]))
        try:
            records.extend(_load_records_from_xlsx(path, province, default_subject_type))
        except Exception as ex:
            logger.warning(f"加载招生数据失败: {path} - {ex}")
    return records


def _resolve_data_path(path_or_url: str) -> Path:
    path = Path(path_or_url)
    if path.is_absolute():
        return path
    return Path(__file__).resolve().parents[1] / path


def _load_records_from_xlsx(
    path: Path,
    province: str,
    default_subject_type: Optional[str],
) -> List[AdmissionRecord]:
    try:
        import openpyxl
    except ImportError as ex:
        raise RuntimeError("读取 .xlsx 招生数据需要安装 openpyxl") from ex

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    major_idx = _header_index(headers, "专业")
    subject_idx = _header_index(headers, "科类")
    year_columns = _year_columns(headers)

    records: List[AdmissionRecord] = []
    for row in rows[1:]:
        major = _cell_text(row, major_idx)
        if not major:
            continue
        subject_type = default_subject_type or _normalize_subject_type(_cell_text(row, subject_idx))
        if not subject_type:
            continue

        for year, cols in year_columns.items():
            score = _to_int(_row_value(row, cols.get("score")))
            rank = _to_int(_row_value(row, cols.get("rank")))
            if score is None or rank is None:
                continue
            records.append(
                AdmissionRecord(
                    year=year,
                    province=province,
                    subject_type=subject_type,
                    major=major,
                    min_score=score,
                    min_rank=rank,
                    plan=None,
                )
            )
    return records


def _header_index(headers: List[str], name: str) -> Optional[int]:
    try:
        return headers.index(name)
    except ValueError:
        return None


def _year_columns(headers: List[str]) -> Dict[int, Dict[str, int]]:
    columns: Dict[int, Dict[str, int]] = {}
    for idx, header in enumerate(headers):
        match = re.match(r"(\d{4})年最低(分|位次)", header)
        if not match:
            continue
        year = int(match.group(1))
        kind = "score" if match.group(2) == "分" else "rank"
        columns.setdefault(year, {})[kind] = idx
    return columns


def _row_value(row: Any, idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _cell_text(row: Any, idx: Optional[int]) -> str:
    value = _row_value(row, idx)
    return str(value).strip() if value is not None else ""


_ADMISSION_RECORDS: List[AdmissionRecord] = []

_MAJOR_ALIASES = {
    "计算机": "计算机科学与技术",
    "计科": "计算机科学与技术",
    "软件": "软件工程",
    "电气": "电气工程及其自动化",
    "电气专业": "电气工程及其自动化",
    "电气工程": "电气工程及其自动化",
    "电气自动化": "电气工程及其自动化",
}

_PROVINCES = [
    "北京", "天津", "河北", "山西", "内蒙古", "辽宁", "吉林", "黑龙江", "上海", "江苏",
    "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南", "广东", "广西",
    "海南", "重庆", "四川", "贵州", "云南", "西藏", "陕西", "甘肃", "青海", "宁夏",
    "新疆",
]

_SUBJECT_TYPES = ["物理类", "历史类", "物理", "历史", "理科", "文科", "综合改革", "综合"]


async def risk_assessment_handler(params: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    MCP Tool handler: async (params, context) -> dict。

    支持两种调用方式：
      1. 结构化参数：province/subject_type/score/rank/major
      2. 原始消息：message="我河南理科 580 分、位次 32000，报计算机稳吗？"
    """
    normalized = _normalize_params(params)
    province = normalized.get("province")
    if province and province not in _SUPPORTED_PROVINCES:
        return _unsupported_province_response(normalized)

    missing = [
        name for name in ("province", "subject_type", "rank", "major")
        if normalized.get(name) in (None, "")
    ]
    if missing:
        return {
            "status": "need_more_info",
            "risk_level": "信息不足",
            "missing": missing,
            "message": "需要补充省份、科类/选科、位次和目标专业后，才能做录取风险判断；如能补充分数，判断会更完整。",
            "params": normalized,
        }

    records = _find_records(
        province=str(normalized["province"]),
        subject_type=str(normalized["subject_type"]),
        major=str(normalized["major"]),
    )
    if not records:
        return {
            "status": "no_data",
            "risk_level": "无法判断",
            "message": "当前已接入的历年录取数据中没有匹配记录，请核对省份、科类/选科、专业名称，或以学校招生网查询结果为准。",
            "params": normalized,
        }

    return _assess(records, normalized)


def _normalize_params(params: Dict[str, Any]) -> Dict[str, Any]:
    message = str(params.get("message") or "")
    extracted = _extract_from_message(message)

    province = params.get("province") or extracted.get("province")
    subject_type = params.get("subject_type") or extracted.get("subject_type")
    major = params.get("major") or extracted.get("major")
    score = params.get("score") or extracted.get("score")
    rank = params.get("rank") or extracted.get("rank")

    if subject_type == "综合":
        subject_type = "综合改革"
    subject_type = _normalize_subject_type(str(subject_type)) if subject_type else None

    if province == "天津" and not subject_type:
        subject_type = "综合改革"

    major = _normalize_major(str(major)) if major else None

    return {
        "province": str(province) if province else None,
        "subject_type": str(subject_type) if subject_type else None,
        "major": major,
        "score": _to_int(score),
        "rank": _to_int(rank),
    }


def _extract_from_message(message: str) -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    for province in _PROVINCES:
        if province in message:
            result["province"] = province
            break

    for subject_type in _SUBJECT_TYPES:
        if subject_type in message:
            result["subject_type"] = subject_type
            break

    for alias, major in _MAJOR_ALIASES.items():
        if alias in message:
            result["major"] = major
            break
    if "major" not in result:
        for record in _ADMISSION_RECORDS:
            if record.major in message:
                result["major"] = record.major
                break

    score_match = re.search(r"(\d{3})\s*分", message)
    if score_match:
        result["score"] = int(score_match.group(1))

    rank_match = re.search(r"(?:位次|排名)\s*[：:是为约大概]*\s*(\d{3,7})", message)
    if rank_match:
        result["rank"] = int(rank_match.group(1))

    return result


def _normalize_major(major: str) -> str:
    major = major.strip()
    return _MAJOR_ALIASES.get(major, major)


def _normalize_subject_type(subject_type: str) -> str:
    value = subject_type.strip()
    mapping = {
        "物理": "物理类",
        "物理类": "物理类",
        "历史": "历史类",
        "历史类": "历史类",
        "理科": "理科",
        "文科": "文科",
        "综合": "综合改革",
        "综合改革": "综合改革",
    }
    return mapping.get(value, value)


def _to_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    match = re.search(r"\d+", str(value))
    return int(match.group(0)) if match else None


_ADMISSION_RECORDS = load_admission_records(_ADMISSION_SOURCES)


def _unsupported_province_response(params: Dict[str, Any]) -> Dict[str, Any]:
    province = params.get("province")
    return {
        "status": "unsupported_province",
        "risk_level": "暂不支持自动判断",
        "province": province,
        "supported_provinces": sorted(_SUPPORTED_PROVINCES),
        "message": (
            f"当前系统只接入了河北、天津的真实历年录取数据，暂不能自动评估{province}考生的录取风险。"
            "建议你在浏览器打开河北工业大学本科招生网或所在省考试院官网，查询该省近三年目标专业的最低位次，"
            "再用自己的位次与近三年最低位次进行对比：位次明显靠前通常更稳，接近最低位次属于冲，"
            "明显靠后则风险较高。"
        ),
        "params": params,
    }


def _find_records(province: str, subject_type: str, major: str) -> List[AdmissionRecord]:
    return sorted(
        [
            record for record in _ADMISSION_RECORDS
            if record.province == province
            and record.subject_type == subject_type
            and record.major == major
        ],
        key=lambda r: r.year,
        reverse=True,
    )


def _assess(records: List[AdmissionRecord], params: Dict[str, Any]) -> Dict[str, Any]:
    score = _to_int(params.get("score"))
    rank = int(params["rank"])
    ranks = [record.min_rank for record in records]
    scores = [record.min_score for record in records]
    avg_rank = int(mean(ranks))
    avg_score = round(mean(scores), 1)
    best_rank_margin = min(record.min_rank - rank for record in records)
    stable_years = sum(
        1
        for record in records
        if rank <= record.min_rank and (score is None or score >= record.min_score)
    )

    if stable_years == len(records) and best_rank_margin >= 1500:
        risk_level = "稳"
        suggestion = "可以作为相对稳妥志愿，但仍建议搭配更稳专业或专业组。"
    elif stable_years >= max(1, len(records) - 1):
        risk_level = "稳中有波动"
        suggestion = "整体有机会，但要关注今年招生计划和报考热度变化。"
    elif stable_years >= 1 or rank <= avg_rank:
        risk_level = "冲"
        suggestion = "有冲刺价值，建议不要只押这一个专业，同时配置更稳的专业。"
    else:
        risk_level = "风险较高"
        suggestion = "从当前已接入数据看位次优势不足，建议作为冲刺项，重点准备更稳和保底方案。"

    return {
        "status": "ok",
        "risk_level": risk_level,
        "province": params["province"],
        "subject_type": params["subject_type"],
        "major": params["major"],
        "user_score": score,
        "user_rank": rank,
        "avg_min_score": avg_score,
        "avg_min_rank": avg_rank,
        "stable_years": stable_years,
        "history": [
            {
                "year": record.year,
                "min_score": record.min_score,
                "min_rank": record.min_rank,
                "plan": record.plan,
            }
            for record in records
        ],
        "reason": _assessment_reason(
            records_count=len(records),
            avg_rank=avg_rank,
            avg_score=avg_score,
            rank=rank,
            score=score,
            stable_years=stable_years,
        ),
        "suggestion": suggestion,
        "disclaimer": "该判断基于当前已接入的历年数据和简单规则，仅供报考参考；最终以河北工业大学本科招生网和省级招生考试机构公布信息为准。",
    }


def _assessment_reason(
    records_count: int,
    avg_rank: int,
    avg_score: float,
    rank: int,
    score: Optional[int],
    stable_years: int,
) -> str:
    if score is None:
        return (
            f"近 {records_count} 年该条件下最低位次均值约为 {avg_rank}，"
            f"最低分均值约为 {avg_score}。你提供的位次为 {rank}，"
            f"其中 {stable_years} 年达到或优于最低位次；因未提供分数，本次主要按位次判断。"
        )
    return (
        f"近 {records_count} 年该条件下最低位次均值约为 {avg_rank}，"
        f"最低分均值约为 {avg_score}。你的位次为 {rank}，"
        f"其中 {stable_years} 年同时达到或优于最低分和最低位次。"
    )


def risk_assessment_fallback(params: Dict[str, Any], context: Any, error: str) -> Dict[str, Any]:
    return {
        "status": "fallback",
        "risk_level": "无法判断",
        "message": "录取风险评估工具暂时不可用，请稍后重试，或以学校招生网和省考试院数据为准。",
        "error": error,
        "params": params,
    }
